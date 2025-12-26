# 📄 backend/services/stock/statuspage_service.py
# 페이지: 재고 현황(StatusPage)
# 역할: 재고현황 조회/스캔/엑셀(xlsx)/재고조정(실사)
# 단계: v1.8 (xlsx 다운로드 전용 export 추가)
#
# ✅ 변경 요약(v1.8)
# - 운영용 xlsx 다운로드: export_operational_xlsx_bytes(sku, selected_skus)
# - 기존 action(export)은 유지(호환) / 프론트는 export-xlsx 사용 권장

from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import anyio
from sqlalchemy import text
from sqlalchemy.orm import Session

from backend.system.error_codes import DomainError

PAGE_ID = "stock.status"
PAGE_VERSION = "v1.8"


async def _run_sync(fn):
    return await anyio.to_thread.run_sync(fn)


async def _execute(session: Session, stmt, params: Optional[Dict[str, Any]] = None):
    params = params or {}
    return await _run_sync(lambda: session.execute(stmt, params))


async def _commit(session: Session):
    return await _run_sync(session.commit)


async def _rollback(session: Session):
    return await _run_sync(session.rollback)


def _safe_user_id(user: Dict[str, Any]) -> str:
    return str(user.get("id") or user.get("user_id") or user.get("username") or "system")


@dataclass
class StatusPageService:
    session: Session
    user: Dict[str, Any]

    # ─────────────────────────────────────────────
    # SQL 베이스
    # ─────────────────────────────────────────────
    def _base_sql_product(self) -> str:
        # 실사용 검색/스캔용: product 기준(원장 없어도 조회 가능)
        return """
        FROM product p
        LEFT JOIN stock_current sc
               ON sc.sku = p.sku
              AND sc.deleted_at IS NULL
        WHERE p.deleted_at IS NULL
          AND p.is_active = TRUE
        """

    def _base_sql_operational(self) -> str:
        # 운영용 리스트: 원장 발생 SKU만
        return """
        FROM (
            SELECT DISTINCT sku
            FROM inventory_ledger
        ) ls
        JOIN product p
          ON p.sku = ls.sku
         AND p.deleted_at IS NULL
         AND p.is_active = TRUE
        LEFT JOIN stock_current sc
               ON sc.sku = p.sku
              AND sc.deleted_at IS NULL
        WHERE 1=1
        """

    # ─────────────────────────────────────────────
    # 0) (호환 유지) 기존 list_items 시그니처
    # ─────────────────────────────────────────────
    async def list_items(
        self,
        *,
        q: Optional[str],
        page: int,
        size: int,
        sort_by: Optional[str] = "sku",
        order: Optional[str] = "asc",
    ) -> Dict[str, Any]:
        return await self.list_operational(
            sku=(q or "").strip() or None,
            page=page,
            size=size,
            sort_by=sort_by,
            order=order,
        )

    # ─────────────────────────────────────────────
    # 1) 운영용 리스트: 원장 발생 SKU만 + SKU 검색만
    # ─────────────────────────────────────────────
    async def list_operational(
        self,
        *,
        sku: Optional[str],
        page: int,
        size: int,
        sort_by: Optional[str] = "sku",
        order: Optional[str] = "asc",
    ) -> Dict[str, Any]:
        if page <= 0 or size <= 0:
            raise DomainError(
                "STOCK-VALID-001",
                detail="page/size 값이 올바르지 않습니다.",
                ctx={"page_id": PAGE_ID, "page": page, "size": size},
            )

        sku_kw = (sku or "").strip()
        where_extra = ""
        params: Dict[str, Any] = {"limit": size, "offset": (page - 1) * size}

        if sku_kw:
            where_extra = " AND (p.sku ILIKE :kw) "
            params["kw"] = f"%{sku_kw}%"

        sort_col = "p.sku"
        if sort_by == "sku":
            sort_col = "p.sku"
        sort_order = "ASC" if (order or "asc").lower() == "asc" else "DESC"

        count_stmt = text(
            f"""
            SELECT COUNT(1) AS cnt
            {self._base_sql_operational()}
            {where_extra}
            """
        )

        list_stmt = text(
            f"""
            SELECT
                p.sku AS sku,
                p.name AS name,
                COALESCE(sc.qty_on_hand, 0) AS current_qty,
                COALESCE(sc.qty_on_hand, 0) - COALESCE(sc.qty_pending_out, 0) AS available_qty,
                sc.last_unit_price AS last_price
            {self._base_sql_operational()}
            {where_extra}
            ORDER BY {sort_col} {sort_order}
            LIMIT :limit OFFSET :offset
            """
        )

        try:
            cnt_res = await _execute(self.session, count_stmt, params)
            count = int((cnt_res.mappings().first() or {}).get("cnt") or 0)

            res = await _execute(self.session, list_stmt, params)
            rows = res.mappings().all()
        except Exception as exc:
            raise DomainError("SYSTEM-DB-901", detail=str(exc), ctx={"page_id": PAGE_ID, "sku": sku_kw})

        items = []
        for r in rows:
            items.append(
                {
                    "sku": r["sku"],
                    "name": r["name"],
                    "current_qty": int(r["current_qty"] or 0),
                    "available_qty": int(r["available_qty"] or 0),
                    "last_price": float(r["last_price"]) if r["last_price"] is not None else None,
                }
            )

        return {"items": items, "count": count, "page": page, "size": size}

    # ─────────────────────────────────────────────
    # 2) 실사용 검색: product 기준 + 상품명/SKU 검색
    # ─────────────────────────────────────────────
    async def search_products(
        self,
        *,
        q: Optional[str],
        page: int,
        size: int,
        sort_by: Optional[str] = "sku",
        order: Optional[str] = "asc",
    ) -> Dict[str, Any]:
        if page <= 0 or size <= 0:
            raise DomainError(
                "STOCK-VALID-001",
                detail="page/size 값이 올바르지 않습니다.",
                ctx={"page_id": PAGE_ID, "page": page, "size": size},
            )

        keyword = (q or "").strip()
        where_extra = ""
        params: Dict[str, Any] = {"limit": size, "offset": (page - 1) * size}

        if keyword:
            where_extra = " AND (p.sku ILIKE :kw OR p.name ILIKE :kw) "
            params["kw"] = f"%{keyword}%"

        sort_col = "p.sku"
        if sort_by == "name":
            sort_col = "p.name"
        elif sort_by == "current_qty":
            sort_col = "COALESCE(sc.qty_on_hand, 0)"
        sort_order = "ASC" if (order or "asc").lower() == "asc" else "DESC"

        count_stmt = text(
            f"""
            SELECT COUNT(1) AS cnt
            {self._base_sql_product()}
            {where_extra}
            """
        )

        list_stmt = text(
            f"""
            SELECT
                p.sku AS sku,
                p.name AS name,
                COALESCE(sc.qty_on_hand, 0) AS current_qty,
                COALESCE(sc.qty_on_hand, 0) - COALESCE(sc.qty_pending_out, 0) AS available_qty,
                sc.last_unit_price AS last_price
            {self._base_sql_product()}
            {where_extra}
            ORDER BY {sort_col} {sort_order}
            LIMIT :limit OFFSET :offset
            """
        )

        try:
            cnt_res = await _execute(self.session, count_stmt, params)
            count = int((cnt_res.mappings().first() or {}).get("cnt") or 0)

            res = await _execute(self.session, list_stmt, params)
            rows = res.mappings().all()
        except Exception as exc:
            raise DomainError("SYSTEM-DB-901", detail=str(exc), ctx={"page_id": PAGE_ID, "q": keyword})

        items = []
        for r in rows:
            items.append(
                {
                    "sku": r["sku"],
                    "name": r["name"],
                    "current_qty": int(r["current_qty"] or 0),
                    "available_qty": int(r["available_qty"] or 0),
                    "last_price": float(r["last_price"]) if r["last_price"] is not None else None,
                }
            )

        return {"items": items, "count": count, "page": page, "size": size}

    # ─────────────────────────────────────────────
    # 2-1) ✅ 운영용 xlsx 다운로드 (bytes 반환)
    #   - 기준: 운영용(원장 발생 SKU만)
    #   - 필터: sku 부분검색 + (옵션) selected_skus 정확일치 목록
    # ─────────────────────────────────────────────
    async def export_operational_xlsx_bytes(
        self,
        *,
        sku: Optional[str] = None,
        selected_skus: Optional[List[str]] = None,
    ) -> Tuple[bytes, str]:
        # openpyxl은 가벼운 편이라 여기서 import
        from io import BytesIO
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        sku_kw = (sku or "").strip()
        clean_skus = [s.strip() for s in (selected_skus or []) if s and s.strip()]
        where_extra = ""
        params: Dict[str, Any] = {}

        if sku_kw:
            where_extra += " AND (p.sku ILIKE :kw) "
            params["kw"] = f"%{sku_kw}%"

        if clean_skus:
            where_extra += " AND (p.sku = ANY(:skus)) "
            params["skus"] = clean_skus

        stmt = text(
            f"""
            SELECT
                p.sku AS sku,
                p.name AS name,
                COALESCE(sc.qty_on_hand, 0) AS current_qty,
                COALESCE(sc.qty_on_hand, 0) - COALESCE(sc.qty_pending_out, 0) AS available_qty,
                sc.last_unit_price AS last_price
            {self._base_sql_operational()}
            {where_extra}
            ORDER BY p.sku ASC
            """
        )

        try:
            res = await _execute(self.session, stmt, params)
            rows = res.mappings().all()
        except Exception as exc:
            raise DomainError("SYSTEM-DB-901", detail=str(exc), ctx={"page_id": PAGE_ID, "sku": sku_kw})

        wb = Workbook()
        ws = wb.active
        ws.title = "stock_status"

        headers = ["SKU", "상품명", "현재수량", "가용수량", "마지막단가"]
        ws.append(headers)

        for r in rows:
            ws.append(
                [
                    r["sku"],
                    r["name"],
                    int(r["current_qty"] or 0),
                    int(r["available_qty"] or 0),
                    float(r["last_price"]) if r["last_price"] is not None else None,
                ]
            )

        # 보기좋게 컬럼 폭 대충 맞춤(과한 계산은 안 함)
        widths = [26, 50, 12, 12, 14]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        bio = BytesIO()
        wb.save(bio)
        content = bio.getvalue()

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"stock_status_{stamp}.xlsx"
        return content, filename

    # ─────────────────────────────────────────────
    # 3) 다건 SKU 조회 (기존 유지: product 기준)
    # ─────────────────────────────────────────────
    async def list_by_skus(
        self,
        *,
        skus: List[str],
        page: int,
        size: int,
        sort_by: Optional[str] = "sku",
        order: Optional[str] = "asc",
    ) -> Dict[str, Any]:
        clean = [s.strip() for s in (skus or []) if s and s.strip()]
        if not clean:
            return {"items": [], "count": 0, "page": page, "size": size}

        sort_col = "p.sku"
        if sort_by == "name":
            sort_col = "p.name"
        elif sort_by == "current_qty":
            sort_col = "COALESCE(sc.qty_on_hand, 0)"
        sort_order = "ASC" if (order or "asc").lower() == "asc" else "DESC"

        stmt = text(
            f"""
            SELECT
                p.sku AS sku,
                p.name AS name,
                COALESCE(sc.qty_on_hand, 0) AS current_qty,
                COALESCE(sc.qty_on_hand, 0) - COALESCE(sc.qty_pending_out, 0) AS available_qty,
                sc.last_unit_price AS last_price
            {self._base_sql_product()}
              AND p.sku = ANY(:skus)
            ORDER BY {sort_col} {sort_order}
            """
        )

        try:
            res = await _execute(self.session, stmt, {"skus": clean})
            rows = res.mappings().all()
        except Exception as exc:
            raise DomainError("SYSTEM-DB-901", detail=str(exc), ctx={"page_id": PAGE_ID})

        items = []
        for r in rows:
            items.append(
                {
                    "sku": r["sku"],
                    "name": r["name"],
                    "current_qty": int(r["current_qty"] or 0),
                    "available_qty": int(r["available_qty"] or 0),
                    "last_price": float(r["last_price"]) if r["last_price"] is not None else None,
                }
            )

        return {"items": items, "count": len(items), "page": page, "size": size}

    # ─────────────────────────────────────────────
    # 4) 엑셀(선택 SKU) (기존 유지: product 기준 / JSON)
    # ─────────────────────────────────────────────
    async def export_items(self, *, selected_skus: List[str]) -> Dict[str, Any]:
        clean = [s.strip() for s in (selected_skus or []) if s and s.strip()]
        if not clean:
            return {"items": [], "count": 0}

        stmt = text(
            f"""
            SELECT
                p.sku AS sku,
                p.name AS name,
                COALESCE(sc.qty_on_hand, 0) AS current_qty,
                COALESCE(sc.qty_on_hand, 0) - COALESCE(sc.qty_pending_out, 0) AS available_qty,
                sc.last_unit_price AS last_price
            {self._base_sql_product()}
              AND p.sku = ANY(:skus)
            ORDER BY p.sku ASC
            """
        )

        try:
            res = await _execute(self.session, stmt, {"skus": clean})
            rows = res.mappings().all()
        except Exception as exc:
            raise DomainError("SYSTEM-DB-901", detail=str(exc), ctx={"page_id": PAGE_ID})

        items = []
        for r in rows:
            items.append(
                {
                    "sku": r["sku"],
                    "name": r["name"],
                    "current_qty": int(r["current_qty"] or 0),
                    "available_qty": int(r["available_qty"] or 0),
                    "last_price": float(r["last_price"]) if r["last_price"] is not None else None,
                }
            )

        return {"items": items, "count": len(items)}

    # ─────────────────────────────────────────────
    # 5) 바코드 스캔 단건 조회 (기존 유지)
    # ─────────────────────────────────────────────
    async def scan_by_barcode(self, *, barcode: str) -> Dict[str, Any]:
        b = (barcode or "").strip()
        if not b:
            raise DomainError("STOCK-VALID-001", detail="barcode가 비어있습니다.", ctx={"page_id": PAGE_ID})

        stmt = text(
            f"""
            SELECT
                p.sku AS sku,
                p.name AS name,
                COALESCE(sc.qty_on_hand, 0) AS current_qty,
                COALESCE(sc.qty_on_hand, 0) - COALESCE(sc.qty_pending_out, 0) AS available_qty,
                sc.last_unit_price AS last_price
            {self._base_sql_product()}
              AND p.barcode = :barcode
            LIMIT 1
            """
        )

        try:
            result = await _execute(self.session, stmt, {"barcode": b})
            row = result.mappings().first()
        except Exception as exc:
            raise DomainError("SYSTEM-DB-901", detail=str(exc), ctx={"page_id": PAGE_ID, "barcode": b})

        if not row:
            raise DomainError(
                "STOCK-NOTFOUND-101",
                detail="해당 바코드로 상품을 찾지 못했어요.",
                ctx={"page_id": PAGE_ID, "barcode": b},
            )

        return {
            "sku": row["sku"],
            "name": row["name"],
            "current_qty": int(row["current_qty"] or 0),
            "available_qty": int(row["available_qty"] or 0),
            "last_price": float(row["last_price"]) if row["last_price"] is not None else None,
        }

    # ─────────────────────────────────────────────
    # 6) 재고 조정(실사) (기존 유지)
    # ─────────────────────────────────────────────
    async def adjust(self, *, payload: Dict[str, Any]) -> Dict[str, Any]:
        sku = (payload.get("sku") or "").strip()
        final_qty = payload.get("final_qty")
        memo_text = (payload.get("memo") or "").strip()

        if not sku:
            raise DomainError("STOCK-VALID-001", detail="sku가 비어있습니다.", ctx={"page_id": PAGE_ID})
        if final_qty is None:
            raise DomainError("STOCK-VALID-001", detail="final_qty가 비어있습니다.", ctx={"page_id": PAGE_ID, "sku": sku})
        try:
            final_qty = int(final_qty)
        except Exception:
            raise DomainError("STOCK-VALID-001", detail="final_qty는 정수여야 합니다.", ctx={"page_id": PAGE_ID, "sku": sku})
        if final_qty < 0:
            raise DomainError("STOCK-VALID-001", detail="final_qty는 0 이상이어야 합니다.", ctx={"page_id": PAGE_ID, "sku": sku, "final_qty": final_qty})

        memo = f"실사조정: {memo_text}" if memo_text else "실사조정"
        actor = _safe_user_id(self.user)

        select_stmt = text(
            """
            SELECT
                id,
                sku,
                qty_on_hand,
                qty_reserved,
                qty_pending_out,
                last_unit_price
            FROM stock_current
            WHERE sku = :sku
              AND deleted_at IS NULL
            FOR UPDATE
            """
        )

        try:
            result = await _execute(self.session, select_stmt, {"sku": sku})
            row = result.mappings().first()

            if row is None:
                product_check_stmt = text(
                    """
                    SELECT sku
                    FROM product
                    WHERE sku = :sku
                      AND deleted_at IS NULL
                      AND is_active = TRUE
                    """
                )
                p_res = await _execute(self.session, product_check_stmt, {"sku": sku})
                product_row = p_res.mappings().first()
                if product_row is None:
                    raise DomainError("STOCK-NOTFOUND-101", detail="해당 SKU의 상품을 찾을 수 없습니다.", ctx={"page_id": PAGE_ID, "sku": sku})

                insert_stock_stmt = text(
                    """
                    INSERT INTO stock_current (
                        sku,
                        qty_on_hand,
                        qty_reserved,
                        qty_pending_out,
                        last_unit_price,
                        total_value,
                        updated_by,
                        updated_at
                    )
                    VALUES (
                        :sku,
                        0,
                        0,
                        0,
                        NULL,
                        NULL,
                        :updated_by,
                        NOW()
                    )
                    ON CONFLICT (sku) DO NOTHING
                    """
                )
                await _execute(self.session, insert_stock_stmt, {"sku": sku, "updated_by": actor})
                await _commit(self.session)

                result = await _execute(self.session, select_stmt, {"sku": sku})
                row = result.mappings().first()

                if row is None:
                    raise DomainError("STOCK-NOTFOUND-101", detail="해당 SKU의 재고 정보를 찾을 수 없습니다.", ctx={"page_id": PAGE_ID, "sku": sku})

        except DomainError:
            raise
        except Exception as exc:
            await _rollback(self.session)
            raise DomainError("SYSTEM-DB-901", detail=str(exc), ctx={"page_id": PAGE_ID, "sku": sku})

        before_qty = int(row["qty_on_hand"] or 0)
        qty_pending_out = int(row["qty_pending_out"] or 0)
        last_unit_price = row["last_unit_price"]

        if final_qty < qty_pending_out:
            raise DomainError(
                "STOCK-STATE-451",
                detail="출고 대기수량보다 작은 값으로 재고를 조정할 수 없습니다.",
                ctx={"page_id": PAGE_ID, "sku": sku, "final_qty": final_qty, "qty_pending_out": qty_pending_out},
            )

        delta = final_qty - before_qty
        qty_in = delta if delta > 0 else 0
        qty_out = -delta if delta < 0 else 0

        try:
            insert_ledger_stmt = text(
                """
                INSERT INTO inventory_ledger (
                    sku,
                    event_type,
                    ref_type,
                    ref_id,
                    qty_in,
                    qty_out,
                    unit_price,
                    memo,
                    created_by,
                    updated_by,
                    created_at
                )
                VALUES (
                    :sku,
                    :event_type,
                    :ref_type,
                    :ref_id,
                    :qty_in,
                    :qty_out,
                    :unit_price,
                    :memo,
                    :created_by,
                    :updated_by,
                    NOW()
                )
                """
            )
            await _execute(
                self.session,
                insert_ledger_stmt,
                {
                    "sku": sku,
                    "event_type": "ADJUST",
                    "ref_type": "STOCK_ADJUST",
                    "ref_id": None,
                    "qty_in": int(qty_in),
                    "qty_out": int(qty_out),
                    "unit_price": last_unit_price,
                    "memo": memo,
                    "created_by": actor,
                    "updated_by": actor,
                },
            )

            update_stock_stmt = text(
                """
                UPDATE stock_current
                SET
                    qty_on_hand = :final_qty,
                    total_value = CASE
                        WHEN last_unit_price IS NULL THEN NULL
                        ELSE (:final_qty * last_unit_price)
                    END,
                    updated_by = :updated_by,
                    updated_at = NOW()
                WHERE sku = :sku
                  AND deleted_at IS NULL
                """
            )
            await _execute(self.session, update_stock_stmt, {"sku": sku, "final_qty": int(final_qty), "updated_by": actor})
            await _commit(self.session)

        except Exception as exc:
            await _rollback(self.session)
            raise DomainError("SYSTEM-DB-901", detail=str(exc), ctx={"page_id": PAGE_ID, "sku": sku})

        after_qty = int(final_qty)
        return {
            "sku": sku,
            "before_qty": before_qty,
            "after_qty": after_qty,
            "current_qty": after_qty,
            "available_qty": after_qty - qty_pending_out,
            "last_price": float(last_unit_price) if last_unit_price is not None else None,
        }
